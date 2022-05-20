from openpyxl import Workbook, load_workbook

wb = load_workbook('监测点A日AQI及首要污染物.xlsx')
ws = wb.active

# fpout = open('监测点A日AQI及首要污染物逐日变化.csv', 'w')
# fpout.write('对比日期, ΔSO2, ΔNO2, ΔPM10, ΔPM2.5, ΔO3, ΔCO, ΔAQI')

# 遍历表格中的数据
start_row = 2
end_row = 820
for row in range(start_row, end_row+1):
    start_col = 3
    end_col = 9
    for cl in range(start_col, end_col+1):
        c = 0
        try:
            c = float(ws.cell(row=row, column=cl).value)
        except:
            legal_prerow = row-1
            legal_nextrow = row+1
            cpre = 0
            cnext = 0
            while True:
                try:
                    cpre = float(ws.cell(row=legal_prerow, column=cl).value)
                    break
                except:
                    if legal_prerow <= start_row:
                        cpre = 0
                        break
                    legal_prerow -= 1
            while True:
                try:
                    cnext = float(ws.cell(row=legal_nextrow, column=cl).value)
                    break
                except:
                    if legal_nextrow >= end_row:
                        cnext = 0
                        break
                    legal_nextrow += 1
            c = (cpre+cnext)/2
        ws.cell(row=row, column=cl).value = c

wb.save('监测点A日AQI及首要污染物-空值处理.xlsx')
