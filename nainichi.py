from openpyxl import load_workbook
from datetime import datetime

wb1 = load_workbook('A.xlsx')
wb2 = load_workbook('监测点A日平均.xlsx')

ws1 = wb1['监测点A逐日污染物浓度实测数据']
ws2 = wb2.active

row1 = 2
row2 = 2

row1max = 820
row2max = 816

# wb1 = load_workbook('BC.xlsx')
# wb2 = load_workbook('监测点B日平均.xlsx')

# ws1 = wb1['监测点B逐日污染物浓度实测数据']
# ws2 = wb2.active

# row1 = 2
# row2 = 2

# row1max = 820
# row2max = 820
fpout = open('A异常日期.txt', 'w', encoding='utf-8')


while row1 < row1max + 1 and row2 < row2max+1:
    # ws2.cell(row2, 2).value 转化为datetime类型
    ws1con = ws1.cell(row1, 1).value
    ws2con = datetime.strptime(ws2.cell(row2, 2).value, "%Y-%m-%d")
    if ws1con != ws2con:
        fpout.write(str(ws1con) + ' <-> '+str(ws2con) + '\n')
        if ws1con < ws2con:
            row1 += 1
        else:
            row2 += 1
    else:
        row1 += 1
        row2 += 1
fpout.close()
