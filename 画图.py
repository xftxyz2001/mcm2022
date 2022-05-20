from openpyxl import load_workbook
import matplotlib.pyplot as plt
plt.rcParams["font.sans-serif"] = ["SimHei"]  # 设置字体
plt.rcParams["axes.unicode_minus"] = False  # 该语句解决图像中的“-”负号的乱码问题


wb = load_workbook('A.xlsx')
ws = wb['监测点A逐日污染物浓度实测数据']

x_date = []
# SO2,NO2,PM10,PM2.5,O3,CO
y_data = {
    'SO2': [],
    'NO2': [],
    'PM10': [],
    'PM2.5': [],
    'O3': [],
    'CO': []
}

keys = ['SO2', 'NO2', 'PM10', 'PM2.5', 'O3', 'CO']
clr = ['b', 'g', 'r', 'c', 'm', 'y']
# 遍历表格中的数据
start_row = 2
end_row = 820
for row in range(start_row, end_row+1):
    x_date.append(ws.cell(row=row, column=1).value.strftime('%Y-%m-%d'))
    start_col = 3
    end_col = 8
    for cl in range(start_col, end_col+1):
        c = 0
        try:
            c = float(ws.cell(row=row, column=cl).value)
        except:
            legal_prerow = row-1
            legal_nextrow = row+1
            cpre = 0
            cnext = 0
            while True:
                try:
                    cpre = float(ws.cell(row=legal_prerow, column=cl).value)
                    break
                except:
                    if legal_prerow <= start_row:
                        cpre = 0
                        break
                    legal_prerow -= 1
            while True:
                try:
                    cnext = float(ws.cell(row=legal_nextrow, column=cl).value)
                    break
                except:
                    if legal_nextrow >= end_row:
                        cnext = 0
                        break
                    legal_nextrow += 1
            c = (cpre+cnext)/2
        y_data[keys[cl-start_col]].append(c)

# print(x_date)
# print(y_data)

for r in range(len(keys)):
    plt.plot(x_date, y_data[keys[r]], clr[r], label=keys[r])
    
plt.title('污染物浓度变化图')
plt.xlabel('日期')  # x轴标题
plt.ylabel('污染物数值')  # y轴标题
plt.legend(keys)
plt.show()

# # 对比两天内同一时刻温度的变化情况
# x = ['5', '8', '12', 'jhfkhjhv', '16', '18', '20']
# y1 = [18, 21, 29, 31, 26, 24, 20]
# y2 = [15, 18, 24, 30, 31, 25, 24]
# # 绘制折线图，添加数据点，设置点的大小
# # * 表示绘制五角星；此处也可以不设置线条颜色，matplotlib会自动为线条添加不同的颜色
# plt.plot(x, y1, 'r', marker='*', markersize=10)
# plt.plot(x, y2, 'b', marker='*', markersize=10)
# plt.title('污染物浓度变化图')  # 折线图标题
# plt.xlabel('时间(h)')  # x轴标题
# plt.ylabel('温度(℃)')  # y轴标题
# # 给图像添加注释，并设置样式
# for a, b in zip(x, y1):
#     plt.text(a, b, b, ha='center', va='bottom', fontsize=10)
# for a, b in zip(x, y2):
#     plt.text(a, b, b, ha='center', va='bottom', fontsize=10)

# # 绘制图例
# plt.legend(['第一天', '第二天'])
# # 显示图像
# plt.show()
