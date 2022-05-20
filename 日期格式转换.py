from openpyxl import load_workbook

# yyyy/mm/dd hh:mm:ss -> yyyy-mm-dd hh:mm:ss

wb = load_workbook('A.xlsx')

ws = wb.active

start_row = 2
end_row = 19433

# print()
# exit(0)
for row in range(start_row, end_row+1):
    ws.cell(row=row, column=1).value = str(ws.cell(row=row, column=1).value)
    print(ws.cell(row=row, column=1).value)

wb.save('A-dateformated.xlsx')
