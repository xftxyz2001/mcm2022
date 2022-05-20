import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# ----------
wb = load_workbook('BC.xlsx')
ws = wb['监测点B逐小时污染物浓度与气象实测数据']

start_row = 2
end_row = 19599

start_col = 9
end_col = 13
# ----------

current_day = ws.cell(start_row, 1).value.date()
current_day_hour = 0
# assert isinstance(current_day, datetime)
# current_day.date()
current_data = {
    '温度(℃)': 0,
    '湿度(%)': 0,
    '气压(MBar)': 0,
    '风速(m/s)': 0,
    '风向(°)': 0
}

# print(current_day)
fpout = open('24avg.csv', 'w')  # gbk
fpout.write('日期,有效时间数,平均温度(℃),平均湿度(%),平均气压(MBar),平均风速(m/s),平均风向(°)\n')
# ----------
fpwarnings = open('监测点B日平均warnings.txt', 'w', encoding='utf-8')
# ----------
for row in range(start_row, end_row+1):
    if ws.cell(row, 1).value.date() != current_day:
        for k, v in current_data.items():
            current_data[k] = current_data[k] / current_day_hour
        s = str(current_day) + ',' + str(current_day_hour) + ',' + str(current_data['温度(℃)']) + ',' + str(current_data['湿度(%)']) + ',' + str(
            current_data['气压(MBar)']) + ',' + str(current_data['风速(m/s)']) + ',' + str(current_data['风向(°)'])
        # print(s)
        fpout.write(s + '\n')
        current_day = ws.cell(row, 1).value.date()
        # print(current_day, current_data, current_day_hour)
        current_data = {
            '温度(℃)': 0,
            '湿度(%)': 0,
            '气压(MBar)': 0,
            '风速(m/s)': 0,
            '风向(°)': 0
        }
        current_day_hour = 1
        continue
    try:
        current_data['温度(℃)'] += float(ws.cell(row, start_col).value)
        current_data['湿度(%)'] += float(ws.cell(row, start_col+1).value)
        current_data['气压(MBar)'] += float(ws.cell(row, start_col+2).value)
        current_data['风速(m/s)'] += float(ws.cell(row, start_col+3).value)
        current_data['风向(°)'] += float(ws.cell(row, start_col+4).value)
        current_day_hour += 1
    except:
        try:
            current_data['温度(℃)'] -= float(ws.cell(row, start_col).value)
            current_data['湿度(%)'] -= float(ws.cell(row, start_col+1).value)
            current_data['气压(MBar)'] -= float(ws.cell(row, start_col+2).value)
            current_data['风速(m/s)'] -= float(ws.cell(row, start_col+3).value)
            current_data['风向(°)'] -= float(ws.cell(row, start_col+4).value)
        except:
            w = '无效数据：'+str(ws.cell(row, 1).value)
            print(w)
            fpwarnings.write(w + '\n')

fpout.close()
fpwarnings.close()

csv = pd.read_csv('24avg.csv', encoding='gbk')
# 把 csv 文件转换为 Excel 文件，并把数据保存在创建的表格 data 中
# ----------
csv.to_excel('监测点B日平均.xlsx', sheet_name='data')
# ----------
