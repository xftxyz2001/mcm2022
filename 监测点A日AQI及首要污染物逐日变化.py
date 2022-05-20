from openpyxl import Workbook, load_workbook
import pandas as pd

wb = load_workbook('监测点A日AQI及首要污染物-空值处理.xlsx')
ws = wb.active

fpout = open('a_aqi_m_daily_change.csv', 'w')
fpout.write('对比日期, ΔSO2, ΔNO2, ΔPM10, ΔPM2.5, ΔO3, ΔCO, ΔAQI')


start_row = 2
end_row = 820
for row in range(start_row, end_row):
    lst = []
    start_col = 3
    end_col = 9
    lst.append(str(ws.cell(row=row+1, column=2).value) +
               '<->' + str(ws.cell(row=row, column=2).value))
    for cl in range(start_col, end_col+1):
        delta = float(ws.cell(row=row+1, column=cl).value) - \
            float(ws.cell(row=row, column=cl).value)
        lst.append(delta)
    print(lst)
    fpout.write('\n' + ','.join(map(str, lst)))

fpout.close()

csv = pd.read_csv('a_aqi_m_daily_change.csv', encoding='gbk')
# 把 csv 文件转换为 Excel 文件，并把数据保存在创建的表格 data 中
csv.to_excel('监测点A日AQI及首要污染物逐日变化.xlsx', sheet_name='data')
