import pymysql
from openpyxl import load_workbook

# wb = load_workbook('BC.xlsx')
wb = load_workbook('监测站A逐小时数据异常值平均值处理后含日期.xlsx')
# 监测站A逐小时数据异常值平均值处理后含日期
# ws = wb['监测点C逐小时污染物浓度与气象实测数据']
ws = wb.active

start_row = 2
end_row = ws.max_row

# 连接数据库
# db = pymysql.connect("localhost", "root", "root", "mcm")
db = pymysql.connect(host='localhost', user='root', password='root', db='mcm')
db.autocommit(True)
# 使用cursor()方法创建一个游标对象
cursor = db.cursor()


def to_float(value):
    try:
        return float(value)
    except:
        return 'NULL'


for row in range(start_row, end_row+1):

    # 使用execute()方法执行SQL语句
    xtime = ws.cell(row=row, column=1).value
    xdd = ws.cell(row=row, column=2).value[-1:]

    so2 = to_float(ws.cell(row=row, column=3).value)
    no2 = to_float(ws.cell(row=row, column=4).value)
    pm10 = to_float(ws.cell(row=row, column=5).value)
    pm25 = to_float(ws.cell(row=row, column=6).value)
    o3 = to_float(ws.cell(row=row, column=7).value)
    co = to_float(ws.cell(row=row, column=8).value)
    wd = to_float(ws.cell(row=row, column=9).value)
    sd = to_float(ws.cell(row=row, column=10).value)
    qy = to_float(ws.cell(row=row, column=11).value)
    fs = to_float(ws.cell(row=row, column=12).value)
    fx = to_float(ws.cell(row=row, column=13).value)
    # sql = f"INSERT INTO hourly VALUES ('{xtime}', '{xdd}', {so2}, {no2}, {pm10}, {pm25}, {o3}, {co}, {wd}, {sd}, {qy}, {fs}, {fx})"
    sql = f"INSERT INTO hourly_copy2 VALUES ('{xtime}', '{xdd}', {so2}, {no2}, {pm10}, {pm25}, {o3}, {co}, {wd}, {sd}, {qy}, {fs}, {fx})"
    print(sql)
    cursor.execute(sql)
    # break

cursor.close()
db.close()
